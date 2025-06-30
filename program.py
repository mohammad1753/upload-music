import os 
import shutil
from PIL import Image
import eyed3
from eyed3.id3.frames import ImageFrame
import re
import openpyxl
import requests
from bs4 import BeautifulSoup
from mutagen.id3 import ID3, COMM
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from tqdm import tqdm
arrays_to_add = []
list_number = 29
### Functioans:

def create_resized_images(image_path, folder_path):
    try:
        # باز کردن تصویر
        img = Image.open(image_path)
        
        # تغییر اندازه و ذخیره اولین عکس (800x800)
        cover_img = img.resize((800, 800))
        cover_path = os.path.join(folder_path, "cover.jpg")
        cover_img.save(cover_path)
        # print(f"تصویر 800x800 به نام cover.jpg در پوشه '{folder_path}' ذخیره شد.")
        
        # تغییر اندازه و ذخیره دومین عکس (300x300)
        m_img = img.resize((300, 300))
        m_path = os.path.join(folder_path, "m.jpg")
        m_img.save(m_path)
        # print(f"تصویر 300x300 به نام m.jpg در پوشه '{folder_path}' ذخیره شد.")
        
        # حذف تصویر اصلی بعد از پردازش
        os.remove(image_path)
        # print(f"تصویر اصلی '{image_path}' حذف شد.")
    except Exception as e:
         print(f"خطا در پردازش تصویر {image_path}: {e}")



def update_music_tags(file_path, song_name, songer_name, cover_path):
    # try:
        name = str(song_name)
        name = name.removeprefix("Ehsan Khajehamiri").replace("(128).mp3","")
        # comment = "[ MahMusic.net ]"
        # بارگذاری فایل MP3
        audio_file = eyed3.load(file_path)
        # حذف اطلاعات قبلی
        
        audio_file.tag.title = None
        audio_file.tag.artist = None
        audio_file.tag.album = None
        audio_file.tag.images.remove
        audio_file.tag.composer = None
        audio_file.tag.album_artist = None
        # تغییر تگ Title
        audio_file.tag.title = f"{name} ~ MahMusic.net"

        # تغییر تگ Artist
        audio_file.tag.artist = f"{songer_name} ~ MahMusic.net"

        # تغییر تگ Album
        audio_file.tag.album = f"{name} (Single)"


        if (audio_file.tag == None):
            audio_file.initTag()
        for i in range(3):
            if audio_file.tag.images:
                audio_file.tag.images.remove

        audio_file.tag.images.set(ImageFrame.FRONT_COVER, open(cover_path,'rb').read(), 'image/jpeg')

        # تغییر تگ Comment


        # ذخیره تغییرات
        audio_file.tag.save()
        
        audio_comm = ID3(file_path)
        audio_comm.add(COMM(encoding=3, lang='eng', desc='', text="[ MahMusic.net ]"))
        audio_comm.save()

        # print(f"تگ‌ها برای {song_name} تغییر کرد.")
        
    # except Exception as e:
    #     print(f"خطا در پردازش فایل {file_path}: {e}")

def get_songtext_site(link):
    response = requests.get(link)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, "html.parser")
        song_text = soup.find("p", style="text-align: center;")
        # print(song_text)

        if song_text:
            # block_text = song_text.get_text(strip=True)
            pattern = r'<p.*?>(.*?)</p>'
            match = re.search(pattern, str(song_text), re.DOTALL)
            if match:
                block_text = match.group(1)
                return block_text

def get_song_name(link, number, c2):
    if c2 == 1:
        response = requests.get(link)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, "html.parser")
            title = soup.find("div",class_="info-sg eng uk-visible@m")
            name = title.h3
            print(name.text)
            return name.text
    elif c2 == 2:
        
        # div1 = driver.find_element(By.ID, "playlist")
        # element = div1.find_element(By.XPATH, f"//div[@data-track='{number}']")
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, f"//div[@data-track='{number}']")))
        div3 = element.find_element(By.CLASS_NAME, "right")
        a = div3.find_element(By.TAG_NAME, "a")
        name_song = a.text
        return name_song



def get_song_link(page_link, name):
    name = name + ".mp3"
    search_box = driver.find_element(By.NAME, "keyword")
    search_box.clear()
    search_box.send_keys(name)
    # element = driver.find_element(By.XPATH, f'//td[@data-sort="{name}"]')
    element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//tr[@style='display: table-row;']")))
    link = element.find_element(By.TAG_NAME, "a")
    link = link.get_attribute("href")
    driver.quit
    print(name ,": Done!")
    return link


def Download_file(Url, Filename, c, number2):
    if c == 1:
        response = requests.get(Url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, "html.parser")
            action = soup.find("div", class_="downloadbutton")
            soup = BeautifulSoup(str(action), "html.parser")
            links = soup.findAll("a", class_="dlbt")
            soup =BeautifulSoup(str(links[0]), "html.parser")
            link = soup.find("a")
            link = link['href']
            
        file_sourc= source_folder +"\\"+ Filename + ".mp3"
        download_with_progress(link, file_sourc)

        # response =requests.get(link)
        # with open(file_sourc, 'wb') as f:
        #     f.write(response.content)
        
        get_image(Url, Filename)
        print(Filename, "Done!")
    elif c == 2:
        # element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, f"//div[@data-track='{number}']")))
        # div3 = element.find_element(By.CLASS_NAME, "left")
        # a = div3.find_elements(By.TAG_NAME, "a")
        # a = a[0]
        # link = a.get_attribute("href")

        file_sourc= source_folder +"\\"+ Filename + ".mp3"
        response =requests.get(Url)
        soup = BeautifulSoup(response.content, 'html.parser')
        if number2 == 0:
            div_elements = soup.find('div', class_='list for-alb active playing')
            link = div_elements.get('data-id')
            print(link)
        else:    
            div_elements = soup.find_all('div', class_='list for-alb')
            for data_link in div_elements:
                soup =BeautifulSoup(str(data_link), 'html.parser')
                d = soup.find("div")
                n = int(d.get('data-track'))
                
                if n == number2:
                    link = d.get('data-id')
                    
                    print(link)
                
            
        
        download_with_progress(link, file_sourc)
        print(Filename, "Done!")

def download_with_progress(url, filename):
    response = requests.get(url, stream=True)
    total_size = int(response.headers.get('content-length', 0))
    with open(filename, 'wb') as f, tqdm(
            desc=filename,
            total=total_size,
            unit='iB',
            unit_scale=True,
            unit_divisor=1024,
        ) as pbar:
        for data in response.iter_content(chunk_size=1024):
            f.write(data)
            pbar.update(len(data))

def get_image(url,filename):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, "html.parser")
    div = soup.find("div", class_="cover-sg")
    soup = BeautifulSoup(str(div), "html.parser")
    img = soup.find("img")
    img_link = img['src']
    img_src= "C:\\Users\\Sarv\\Desktop\\upload-music\\New folder2" +"\\"+ filename + ".jpg"
    response =requests.get(img_link)
    with open(img_src, 'wb') as f:
        f.write(response.content)



def get_infosong(link):
    response = requests.get(link)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, "html.parser")
        info_text = soup.find("div", class_="info-sg")
        
        soup2 = BeautifulSoup(str(info_text), "html.parser")
        tags = soup2.find_all("p")
        extracted_text = ""
        
        for tag in tags:
            for i_tag in tag.find_all("i"):
                i_tag.decompose()
            # q = str(tag.text)
            extracted_text = str(extracted_text+ tag.text + '<br/>')
        if extracted_text != "":
            # text = str(extracted_text)
            result = extracted_text.strip()
            # print("infooooo", result)

            return result
        else:
            return None
                
            
             
def song_persian_name(link):
    response = requests.get(link)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, "html.parser")
        info_text = soup.find_all("div", class_="info-sg")
        
        soup2 = BeautifulSoup(str(info_text[1]), "html.parser")
        blocktext = soup2.find("h3")
        pattern = r"<h3>(.*?)</h3>"
        match = re.search(pattern, str(blocktext), re.DOTALL)
        if match:
            extracted_text = match.group(1).strip()
        return extracted_text
            

# href="/dl/majid-razavi/Garsha%20Rezaei%20-%20Aksaye%20Akhar%20%28128%29.mp3"

def get_info_exel_album(folders, source_folder, arrays_to_add, min_row, number):
    link_list = []
    for row in sheet.iter_rows(min_row, values_only=True):  # شروع از ردیف 2 برای صرف نظر از هدر
        print(number)
        song_english_name = row[0]
        if song_english_name is None:
            break  
        song_english_name = song_english_name.replace(" - ", " ")
        page_link_playmusic = row[1]
        song_link = row[2]
        song_persian_names = row[5]
        lyric_song = row[4]
        if lyric_song is None:
            lyric = ""
        else: 
            lyric = lyric_song.replace("\n\n", "<br/>").replace("\n", "<br/>")
        pattern = r"(?<=https://mahmusic.net/dl/).*?(?=\.mp3)"
        match = re.search(pattern, song_link)
        print("@song english name: ", song_english_name)
        if match:
            result = match.group()
            # print(result)

        for folder in folders:
            song_name =folder
            song_folder_name = str(change_music_name(song_name))
            song_name = str(song_folder_name.replace("-", " ").title())

            if song_english_name == song_name:
                info = get_infosong(page_link_playmusic)
                if info is None:
                    info = ""
                values_to_update = {
                        "nn3": song_persian_names,
                        "nn13": song_name,
                        "nn4": song_folder_name,
                        "nn5": result,
                        "nn10": lyric,
                        "nn11": info
                        }
                # print(values_to_update, "\n \n")
                    # جستجوی فایل php در پوشه مربوطه
                folder_path = os.path.join(source_folder, song_folder_name)
                php_file_path = os.path.join(folder_path, "music.php")
                if os.path.exists(php_file_path):
                    
                    link_list.append(update_music_php_file(php_file_path, values_to_update,arrays_to_add, number))
                    number = number + 1
    return link_list

def get_info_exel_single(folders, source_folder, arrays_to_add, number, min_row=2):
    link_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # شروع از ردیف 2 برای صرف نظر از هدر
        print(number)
        song_english_name = row[0]     # اسم آهنگ
        page_link_playmusicsite = row[1]     
        song_link = row[2]
        if song_link is not None:
            pattern = r"(?<=https://mahmusic.net/dl/).*?(?=\.mp3)"
            match = re.search(pattern, song_link)
        else:
            break
        if match:
            result = match.group()

        for folder in folders:
                song_name =folder
                song_folder_name = str(change_music_name(song_name))
                song_name = str(song_folder_name.replace("-", " ").title())    
                # ساخت دیکشنری مقادیر برای متغیرها
                if song_english_name == song_name:
                    print("@song english name: ", song_english_name)

                    info = get_infosong(page_link_playmusicsite)
                    
                    if info is None:
                        info = ""
                    Lyrics = get_songtext_site(page_link_playmusicsite)
                    if Lyrics is None:
                        Lyrics = ""
                    values_to_update = {
                        "nn3": song_persian_name(page_link_playmusicsite),
                        "nn13": song_name,
                        "nn4": song_folder_name,
                        "nn5": result,
                        "nn6": result,
                        "nn10": Lyrics,
                        "nn11": info
                        }
                    # جستجوی فایل php در پوشه مربوطه
                    folder_path = os.path.join(source_folder, song_folder_name)
                    php_file_path = os.path.join(folder_path, "music.php")
                    if os.path.exists(php_file_path):
                        
                        link_list.append(update_music_php_file(php_file_path, values_to_update,arrays_to_add, number))
                        number = number + 1
    return link_list
    
def get_list_link(link_list):
    if link_list is None:
        print("none")
    else:
        for link in link_list:
            with open("C:\\Users\\Sarv\\Desktop\\upload-music\\aron-afshar\\list.txt", "+a", encoding="utf-8") as file:
                file.write(link)


        
def update_file_php(arrays_to_add):
    if os.path.exists(output_php_file):
        with open(output_php_file, 'r', encoding='utf-8') as php_file:
            php_content = php_file.read()
        
        # پیدا کردن آرایه $listasli
        match = re.search(r'\$listasli\s*=\s*array\((.*)\);\s*', php_content, re.DOTALL)
        if match:
            # استخراج محتوای قبلی آرایه
            existing_content = match.group(1).strip()
            new_content = existing_content + "\n".join(arrays_to_add)
            
            # بازنویسی آرایه $listasli با محتوای جدید
            updated_php_content = re.sub(r'\$listasli\s*=\s*array\((.*)\);\s*',f"$listasli = \n array({new_content});",php_content,flags=re.DOTALL)
            php_file_text = updated_php_content.splitlines()
            updated_php_content = convertListToString(php_file_text)
            target_word = "array"
            updated_text = re.sub(rf"({target_word})", r"\n\1", updated_php_content)
            # print(updated_text)

            # نوشتن محتوا به فایل PHP
            with open(output_php_file, 'w', encoding='utf-8') as php_file:
                php_file.write(updated_text)
        else:
            print("آرایه $listasli در فایل PHP یافت نشد!")
    else:
        print(f"فایل {output_php_file} یافت نشد. ایجاد فایل جدید...")
        with open(output_php_file, 'w', encoding='utf-8') as php_file:
            php_file.write("<?php\n\n")
            php_file.write("$listasli = array(\n")
            php_file.writelines(arrays_to_add)
            php_file.write(");\n")
            php_file.write("\n?>")
    # print("okkk!")

    arrays_to_add = []

def update_music_php_file(file_path, values,arrays_to_add, number):
    """
    تغییر متغیرهای مشخص شده در فایل PHP.
    
    :param file_path: مسیر فایل PHP.
    :param values: دیکشنری حاوی مقادیر جدید برای متغیرها.
    """
    # try:
        # باز کردن فایل برای خواندن
    with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        
        # تغییر متغیرهای موردنظر
    for key, value in values.items():
            # value = str(value)
            
            if key == "nn11":
                # if value:
                #     print(value)
                # else:
                #     print("nooooooot valueee")
                pattern = r'\$nn11\s*=\s*"<strong>\$nn13<\/strong>(<br\s*\/>)(?=")'
                replacement =  rf'\g<0>{value}'
                content = re.sub(pattern, replacement, content)
                match = re.search(pattern, content)
                
                # if match:
                #      print("yessss")
                # else:
                #      print("nooooo")
            else:
                pattern = rf"(\${key}\s*=\s*['\"])(.*?)"
                replacement = rf'\g<0>{value}'
                content = re.sub(pattern, replacement, content)
        # ذخیره فایل با تغییرات
    with open(file_path, 'w', encoding='utf-8') as file:
            file.write(content)
        
    # print(f"تغییرات در فایل {file_path} با موفقیت اعمال شد.")
    # except Exception as e:
    # print(f"خطا در پردازش فایل {file_path}: {e}")
    with open(file_path, 'r', encoding='utf-8') as php_file:
        php_content = php_file.read()
        # جستجو برای نام پوشه و لینک موزیک در فایل PHP
        folder_name_in_php = re.search('\$nn4\s*=\s*["\'](.*?)["\'];', php_content)
        link_in_php = re.search('\$nn5\s*=\s*["\'](.*?)["\'];', php_content)
        # link_in_php = str(link_in_php).replace("<re.Match object; span=(1021, 1717), match='$nn5=", "")
        # print("??????? ", link_in_php)

        song_name_in_php = re.search('\$nn3\s*=\s*["\'](.*?)["\'];', php_content)
        if folder_name_in_php and link_in_php and song_name_in_php:
                    folder_name_from_php = folder_name_in_php.group(1)  # استخراج اسم پوشه از فایل music.php
                    link_from_php = link_in_php.group(1)  # استخراج لینک موزیک از فایل music.php
                    song_name_from_php = song_name_in_php.group(1)  # استخراج اسم آهنگ از فایل music.php
                    
                    # ساخت آرایه برای آهنگ
                    song_array = f"array('file' => '{link_from_php}','artist' => 'آرون افشار','namemusic' => '{song_name_from_php}', 'cover' => '/list/aron-afshar/{folder_name_from_php}/cover.jpg', 'nns' => '{folder_name_from_php}','link' => '/list/aron-afshar/{folder_name_from_php}/'),"                    
                    # arrays_to_add.append(song_array)              
                    # song_array = song_array.split()
                    # php_file_text = convertListToString(song_array)
                    
                    link_list = ""
                    link_list = str (number) + ". " + f"https://mahmusic.net/list/aron-afshar/{folder_name_from_php}/\n"
                    update_file_php(song_array)
                    
                    return link_list

                    
                    

def convertListToString(myList):
    return "".join(map(str, myList))

    
def change_music_name(string):
    string = str(string)
    string = string.removeprefix("Aron Afshar - ")
    string = string.replace(" (128)", "")
    string = string.replace(".mp3", "")
    string = string.replace(".jpg", "")
    string = string.replace(" (Ft Ehsan Khajehamiri)", "").replace("Benyamin & Ehsan Khajehamiri - ", "").replace(" (6&8 Remix By Dj Mamas)", "").replace(" (Ft Amin Ghazi)", "")
    string = string.replace("Dj Ramin - ", "").replace("Dj Philip - ", "").replace(" (Top30D.Biz)","")
    string = string.replace("Ehsan Khajehamiri & Afshin Zarei - ","").replace("Ehsan Khajehamiri & Amin Bamshad - ","").replace("Ehsan Khajehamiri & Mohammad Reza Moshiri - ","").replace("Ehsan Khajehamiri Ft Amir Irannejad - ","").replace("Ehsan Khajehamiri Ft Mohammad Mojerloo - ","")
    string = string.replace (" - ", "-")
    string = string.replace("   ", " ").replace("  ", " ")
    string = string.replace(" ", "-").lower()
    return string

# مسیر پوشه‌ای که فایل‌های MP3 در آن قرار دارند
output_php_file = "C:\\Users\\Sarv\\Desktop\\upload-music\\aron-afshar\\file.php" # فایل PHP کلی که آرایه‌ها باید به آن اضافه شوند
# excel_file = "C:\\Users\\Sarv\\Desktop\\upload-music\\aron-afshar\\aron-afshar.xlsx"  # مسیر فایل اکسل
excel_file = "C:\\Users\\Sarv\\Desktop\\upload-music\\aron-afshar\\sample.xlsx"  # مسیر فایل اکسلپ
source_folder = "C:\\Users\\Sarv\\Desktop\\upload-music\\New-folder"
name_songer = "Aron Afshar"
min_row = 1
link_numb = 1
# دریافت لیست پوشه‌ها در پوشه اصلی
folders = os.listdir(source_folder)
# خواندن اطلاعات از فایل اکسل
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active  # فرض می‌کنیم اطلاعات در شیت اول باشند
# دریافت لیست فایل‌ها در پوشه
files = os.listdir(source_folder)
folders = os.listdir(source_folder)
ch = 1
while ch != 0:
    print("""pls chose one case
          1.Send music files to special folder
          2.check all folders and change info
          3.creat a folder with mp3 files
          4.Info in exel file
          5.Download files
          6.Upload files
          7.Send imgs into folders
          0.Exit""")
    ch =    int(input("Enter your choose: "))
    if ch == 1 :
        c = 1
        print("""choose a number:
                1.my musics are single
                2.my musics are album""")
        c = int(input("Enter number: "))
        if c == 1:
            for file in files:
            # بررسی اینکه فایل پسوند .mp3 دارد
                if file.endswith(".mp3"):
                # حذف پسوند و جایگزینی فاصله‌ها با خط تیره
                    folder_name = change_music_name(file)
                # مسیر کامل پوشه جدید
                    new_folder_path = os.path.join(source_folder, folder_name)    
                # folder_names_list.append(folder_name)
                    os.makedirs(new_folder_path)
                
                # انتقال فایل به داخل پوشه مربوطه
                    source_file_path = os.path.join(source_folder, file)  # مسیر فایل اصلی
                    destination_file_path = os.path.join(new_folder_path, file)  # مسیر مقصد
                    shutil.move(source_file_path, destination_file_path)
                
                    img_folder = "C:\\Users\\Sarv\\Desktop\\upload-music\\New folder2"
                    imgs = os.listdir(img_folder)
                    for img in imgs :
                        img = str(img).replace(".jpg", "").lower()
                        file = file.replace(".mp3", "").lower()
                        if img == file:
                            img = str(img) +".jpg"
                            img_path = img_folder + "\\" + img 
                            f_path = source_folder + "\\" +file + ".jpg"
                            shutil.copy(img_path, f_path)
        if c == 2:
            for file in files:
            # بررسی اینکه فایل پسوند .mp3 دارد
                if file.endswith(".mp3"):
                # حذف پسوند و جایگزینی فاصله‌ها با خط تیره
                    folder_name = change_music_name(file)
                # مسیر کامل پوشه جدید
                    new_folder_path = os.path.join(source_folder, folder_name)    
                # folder_names_list.append(folder_name)
                    os.makedirs(new_folder_path)
                
                # انتقال فایل به داخل پوشه مربوطه
                    source_file_path = os.path.join(source_folder, file)  # مسیر فایل اصلی
                    destination_file_path = os.path.join(new_folder_path, file)  # مسیر مقصد
                    shutil.move(source_file_path, destination_file_path)    

    if ch == 7:
        c = 1
        while c != 0:
            print("""Choose a Number 
            1.Singel Music
            2.Album Music
            0.Back""") 
            c = int(input("Enter a Number: "))
            if c == 1:
                for file in files: 
                    if file.endswith(".jpg") :
                        l = file.replace(" .jpg", "").replace(".jpg", "").replace(" ", "-").lower()
                        print(l)
                        f_path = source_folder + "\\" + l +"\\"+ file
                        img_path = source_folder + "\\" + file
                        shutil.move(img_path, f_path)
            elif c == 2:
                for file in files: 
                    if file.endswith(".jpg"):
                        
                        pattern = r"\s+.*"
                        match= re.sub(pattern, "", file)                       
                        l = str(file).replace(" .jpg", "").replace(".jpg", "").replace(" ", "-").lower()                        
                        l = str(file).replace(" .jpg", "").replace(".jpg", "").replace(" ", "-").lower()
                            
                        print(l)
                        f_path = source_folder + "\\" + l +"\\"+ str(file)
                        img_path = source_folder + "\\" + str(file)
                        shutil.move(img_path, f_path)
    if ch == 4:
        c = 1
        while c != 0 :
            print("""Choose a number: 
            1.Insert Name In Excel File
            2.Insert Music Link in Excel File
            0.Back""")
            c = int(input("Enter a Number: "))
            if c == 1:
                c2 = 1
                while c2 != 0 :
                    print("""Choose a Number:
                        1.Single Music
                        2.Album Music""")
                    c2 = int(input("Enter Number: "))
                    if c2 == 1:
                        name_list = []
                        for row in sheet.iter_rows(min_row = 2 , values_only=True):  # شروع از ردیف 2 برای صرف نظر از هدر
                            page_link = row[1]
                            name_list.append(get_song_name(page_link, None, c2))
                        for row in range(1, len(name_list) + 1):
                            sheet.cell(row=row+1, column=1, value=name_list[row - 1])
                        workbook.save(excel_file)
                    if c2 == 2:
                        name_list = []
                        driver = webdriver.Firefox()
                        for row in sheet.iter_rows(min_row = 2 , values_only=True):  # شروع از ردیف 2 برای صرف نظر از هدر
                            page_link = row[1]
                            number = row[3]
                            if number == 0:
                                driver.get(page_link)
                            name_list.append(get_song_name(page_link, number, c2))
                        driver.close()
                        
                        for row in range(1, len(name_list) + 1):
                            sheet.cell(row=row+1, column=1, value=name_list[row - 1])
                        workbook.save(excel_file)   
            elif c == 2:
                link_list = []
                songer_name = name_songer.replace(" ", "-").lower()
                page_links = "https://mahmusic.net/dl/" + str(songer_name)
                driver = webdriver.Firefox()
                driver.get(str(page_links))
                
                for row in sheet.iter_rows(min_row = 2, values_only = True ):
                    name = row[0]
                    if name is not None:
                        link_list.append(get_song_link(page_links, name))
                    else:
                        link_list.append("none")
                for row in range(1, len(link_list) + 1):
                    sheet.cell(row=row+1, column=3, value=link_list[row - 1])
                workbook.save(excel_file) 
                
        
    if ch == 5:
        c = 1
        while c != 0:
            print("""Choose a Number:
            1.Single Music 
            2.Album Music
            0.Back""")
            c = int(input("Enter a Number: "))
            if c == 1:
                for row in sheet.iter_rows(min_row = 2 , values_only=True):
                    name = row[0]
                    page_link = row[1]
                    Download_file(page_link, name, c, None)
            if c == 2:
                
                for row in sheet.iter_rows(min_row = 2 , values_only=True):
                    name = row[0]
                    page_link = row[1]
                    number2 = int(row[3])
                    number2 = number2 -1
                    Download_file(page_link, name, c, number2)




    if ch == 6:
        for folder in folders:
            folder_path = os.path.join(source_folder, folder)
            for file in os.listdir(folder_path):
                if file.endswith(".mp3"):
                    file_path = os.path.join(folder_path, file)
                    files = {f"{file}": open(file_path, 'rb')}
                    username = input("Enter Username : ")
                    Pass = input("Enter Pass : ")
                    data = {'username': username, 'password': Pass}
                    url = "https://mahmusic.net/dl/hoorosh-band/"
                    response = requests.post(url, files=files, data=data)
                    if response.status_code == 200:
                        print("فایل با موفقیت آپلود شد.")
                    else:
                        print("خطا در آپلود فایل:", response.text)




    if ch == 3:
        new_folder_path = os.path.join(source_folder, "all_musics")
        os.makedirs(new_folder_path)
        for folder in folders:
            folder_path = os.path.join(source_folder, folder)
            for file in os.listdir(folder_path): 
                if file.endswith(".mp3"):
                    file_path3 = source_folder+ "\\"+ folder+ "\\"+ file
                    shutil.copy(file_path3, new_folder_path)
                            
            
                
    if ch==2:
        ch2 = 1
        while ch2 != 0:
            min_row = sheet.min_row + 1#یکی اضاف کردم برای babak jahanbakhsh
            print("""choose one case:
                        1.Edit MP3 files'Tags
                        2.Edit music.php
                        3.Edit pictures size
                        4.Insert Index.php
                        5.Delete mp3 files
                        0.Exit""")
            ch2 = int(input("Enter your choose: "))
            if ch2 == 2:
                ch3 = 1
                while ch3 != 0:
                    print("""chose a number:
                                        1.single music
                                        2.album music
                                        0.Exit""")
                    ch3 = int(input("Enter your choice: "))
                    if ch3 == 1:
                        # for i in range(sheet.max_row + 1):
                            string = get_info_exel_single(folders, source_folder, arrays_to_add, list_number)
                            min_row += 1
                            get_list_link(string)
                    elif ch3 == 2:
                        #  for i in range(sheet.max_row + 1):
                            string = get_info_exel_album(folders, source_folder, arrays_to_add, min_row, list_number)
                            min_row += 1
                            get_list_link(string)
            for folder in folders:
                # بررسی اینکه آیا این یک پوشه است
                min_row += 1
                folder_path = os.path.join(source_folder, folder)
                if os.path.isdir(folder_path):
                    if ch2 == 4:
                        index_php_path = "C:\\Users\\Sarv\\Desktop\\upload-music\\index.php"  
                        S_folder_path = folder_path 
                        shutil.copy(index_php_path, S_folder_path) 
                        music_php_path = "C:\\Users\\Sarv\\Desktop\\upload-music\\music.php"  
                        S_folder_path = folder_path
                        shutil.copy(music_php_path, S_folder_path) 

                    # جستجو برای فایل‌های تصویری در هر پوشه
                    for file in os.listdir(folder_path):
                        # print("file: ", file)
                        song_name =folder
                        song_folder_name = str(change_music_name(song_name))
                        song_name = str(song_folder_name.replace("-", " ").title())
                            
                        if ch2 == 1 :
                        
                                if file.endswith(".mp3"):
                                    file_tag_path = source_folder + "\\" +str(change_music_name(file))+ "\\" + file
                                    
                                    cover_path = source_folder + "\\" +str(change_music_name(file))+ "\\" + "cover.jpg"
                                    # print("path of file:", file_tag_path )
                                    # print("music name: ", file_tag_path)
                                    update_music_tags(file_tag_path, song_name,name_songer, cover_path)
                            
                        # if ch2 == 4:
                        #     if file=="music.php":
                        #         file_php_path = source_folder + "\\" +str(change_music_name(file)) + "\\" + "music.php"
                        #         # print(song_name)
                                
                                # get_list_link(link_list)
                        if ch2 == 3:
                            # بررسی اینکه آیا فایل پسوند .jpg یا .png دارد
                            if file.lower().endswith((".jpg", ".jpeg", ".png")):
                                image_path = os.path.join(folder_path, file)
                                
                                create_resized_images(image_path, folder_path)
                        if ch2 == 5:
                            if file.endswith(".mp3"):
                                file_p = source_folder+ "\\" + folder+ "\\" + file
                                os.remove(file_p)
                                # file_p2 = source_folder+ "\\" + folder+ "\\cover.jpg"
                                # os.remove(file_p2)
                        # print("تمام تصاویر با موفقیت تغییر اندازه یافتند!")
                # if ch2 == 4:
                #     php_file__path = get_info_exel(song_name, song_folder_name)
                #     get_info_from_musicphp(php_file__path,arrays_to_add)